# ============================================================
# PANEL ADMIN — ÉTAPE 8 : FICHIER EXCEL AVEC SOMMAIRE STRUCTURÉ
# ============================================================
# Lit le fichier d'indicateurs produit par 07_calcul_indicateur.py
# et produit un fichier Excel navigable avec un sommaire à 4 sections.
#
# Structure du fichier produit :
#   FEUILLE SOMMAIRE (liens cliquables) :
#     Section 1 : Classification par grand groupe CITP (9 groupes)
#     Section 2 : Classification par grade (A1-D3)
#     Section 3 : Classification par sexe (Homme / Femme)
#     Section 4 : Multi-postes par grade
#
#   Feuilles de détail (une par entrée du sommaire) :
#     CITP_GG1 … CITP_GG9  — données CITP filtrées par grand groupe
#     GRADE_A1 … GRADE_D3   — indicateurs filtrés par grade
#     SEXE_Homme / SEXE_Femme
#     MULTI_A1 … MULTI_D3   — multi-postes filtrés par grade
#
# Source  : s3://staging/panel_admin/exports_gold/indicateurs_*.xlsx
# Sortie  : s3://staging/panel_admin/exports_gold/indicateurs_*_SOMMAIRE.xlsx
#
# ⚠️  IMPORTANT : les 3 paramètres ci-dessous doivent être identiques
#     à ceux utilisés lors de l'exécution de 07_calcul_indicateur.py.
#
# Dépendances :
#   pip install polars boto3 python-dotenv openpyxl
# ============================================================

import io
import os
import re
from dotenv import load_dotenv
import boto3
from botocore.client import Config
import polars as pl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

load_dotenv(".env")

# ============================================================
# PARAMÈTRES — doivent correspondre exactement à ceux de
# 07_calcul_indicateur.py pour retrouver le bon fichier source
# ============================================================

MODE_SALAIRE_BRUT  = True   # True = salaire_brut | False = revenu_salarial
PERIODE_RECENTE    = True   # True = 2024-2025 | False = 2015 → récent
INCLURE_ZEROS      = True   # True = avec zéros | False = sans zéros

# ============================================================

# LORSQU'ON TRAVAILLE DEPUIS SA MACHINE LOCAL
MINIO_ENDPOINT   = "http://192.168.1.230:30137"
MINIO_ACCESS_KEY = "datalab-team"
MINIO_SECRET_KEY = "minio-datalabteam123"

# LORSQU'ON TRAVAILLE SUR JHUB
# MINIO_ENDPOINT   = "http://minio.mon-namespace.svc.cluster.local:80"
# MINIO_ACCESS_KEY = "datalab-team"
# MINIO_SECRET_KEY = "minio-datalabteam123"

BUCKET_STAGING = "staging"
PREFIX_EXPORTS = "panel_admin/exports_gold"

_mode_label    = "salaire_brut"  if MODE_SALAIRE_BRUT else "revenu_salarial"
_periode_label = "2024_2025"     if PERIODE_RECENTE   else "2015_recent"
_zeros_label   = "avec_zeros"    if INCLURE_ZEROS     else "sans_zeros"
_base_name     = f"indicateurs_{_mode_label}_{_periode_label}_{_zeros_label}"
INPUT_KEY      = f"{PREFIX_EXPORTS}/{_base_name}.xlsx"
OUTPUT_KEY     = f"{PREFIX_EXPORTS}/{_base_name}_SOMMAIRE.xlsx"

# Classification CITP grands groupes
CITP_GRANDS_GROUPES = {
    1: "1 - Directeurs, cadres de direction et gérants",
    2: "2 - Professions intellectuelles et scientifiques",
    3: "3 - Professions intermédiaires",
    4: "4 - Employés de type administratif",
    5: "5 - Personnel des services directs aux particuliers, commerçants et vendeurs",
    6: "6 - Agriculteurs et ouvriers qualifiés de l'agriculture, de la sylviculture et de la pêche",
    7: "7 - Métiers qualifiés de l'industrie et de l'artisanat",
    8: "8 - Conducteurs d'installations et de machines, et ouvriers de l'assemblage",
    9: "9 - Professions élémentaires",
}

# --- CLIENT S3 ---
s3 = boto3.client(
    "s3",
    endpoint_url          = MINIO_ENDPOINT,
    aws_access_key_id     = MINIO_ACCESS_KEY,
    aws_secret_access_key = MINIO_SECRET_KEY,
    config                = Config(
        signature_version = "s3v4",
        retries           = {"max_attempts": 5, "mode": "adaptive"},
        connect_timeout   = 30,
        read_timeout      = 120,
    ),
    region_name = "us-east-1",
    verify      = False,
)

print("\n" + "=" * 70)
print("CRÉATION FICHIER AVEC SOMMAIRE STRUCTURÉ (4 SECTIONS)")
print("=" * 70 + "\n")

# ============================================================
# VÉRIFICATION ET CHARGEMENT
# ============================================================

print("✓ Chargement du fichier source...")
print(f"  s3://{BUCKET_STAGING}/{INPUT_KEY}\n")

buf = io.BytesIO()
try:
    s3.download_fileobj(BUCKET_STAGING, INPUT_KEY, buf)
except Exception as e:
    raise FileNotFoundError(
        f"Fichier introuvable : s3://{BUCKET_STAGING}/{INPUT_KEY}\n"
        f"Vérifiez que les paramètres MODE_SALAIRE_BRUT / PERIODE_RECENTE / INCLURE_ZEROS\n"
        f"correspondent exactement à ceux utilisés dans 07_calcul_indicateur.py\n{e}"
    )
buf.seek(0)

# Listing des feuilles disponibles
wb_source = load_workbook(buf, read_only=True, data_only=True)
feuilles_existantes = wb_source.sheetnames
print(f"Feuilles disponibles ({len(feuilles_existantes)}) :")
for f in feuilles_existantes:
    print(f"  • {f}")
print()
wb_source.close()

# Lecture des données avec Polars
buf.seek(0)
xlsx_bytes_source = buf.getvalue()


def lire_feuille(nom: str) -> pl.DataFrame | None:
    if nom not in feuilles_existantes:
        return None
    buf2 = io.BytesIO(xlsx_bytes_source)
    try:
        df = pl.read_excel(buf2, sheet_name=nom)
        print(f"  ✓ {nom} : {len(df)} lignes")
        return df
    except Exception as e:
        print(f"  ✗ {nom} : {e}")
        return None


print("📖 Lecture des feuilles...")
dt_citp          = lire_feuille("REVENU_CITP_Detail")
dt_grade         = lire_feuille("REVENU_Grade_Detail")
dt_grade_sexe    = lire_feuille("REVENU_Grade_Sexe")
dt_multi         = lire_feuille("MULTI_Par_Grade_NbPostes")
dt_bareme        = lire_feuille("REVENU_Bareme_Detail")
dt_bareme_grade  = lire_feuille("REVENU_Bareme_Grade")
dt_bareme_sexe   = lire_feuille("REVENU_Bareme_Sexe")
dt_modele_anstat = lire_feuille("MODELE_ANSTAT")
BAREME_DISPONIBLE = dt_bareme is not None
print()

# ============================================================
# FONCTIONS UTILITAIRES EXCEL
# ============================================================

FILL_TITLE    = PatternFill("solid", fgColor="1F4E78")
FILL_SUBTITLE = PatternFill("solid", fgColor="4472C4")
FILL_HEADER   = PatternFill("solid", fgColor="4F81BD")

FONT_TITLE    = Font(color="FFFFFF", bold=True, size=16)
FONT_SUBTITLE = Font(color="FFFFFF", bold=True, size=14)
FONT_HEADER   = Font(color="FFFFFF", bold=True, size=11)
FONT_LINK     = Font(color="0563C1", underline="single", size=12)

ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center")


def style_cell(cell, fill=None, font=None, align=None):
    if fill:  cell.fill      = fill
    if font:  cell.font      = font
    if align: cell.alignment = align


def merge_titre(ws, texte: str, row: int, n_cols: int = 4,
                fill=FILL_TITLE, font=FONT_TITLE) -> None:
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=n_cols
    )
    cell = ws.cell(row=row, column=1, value=texte)
    style_cell(cell, fill=fill, font=font, align=ALIGN_CENTER)
    ws.row_dimensions[row].height = 28


def lien_interne(ws, row: int, col: int, texte: str, feuille_cible: str) -> None:
    feuille_safe = feuille_cible.replace("'", "''")
    cell = ws.cell(row=row, column=col,
                   value=f'=HYPERLINK("#{feuille_safe}!A1","{texte}")')
    style_cell(cell, font=FONT_LINK, align=ALIGN_LEFT)


def ecrire_entetes(ws, entetes: list[str], row: int) -> None:
    for j, h in enumerate(entetes, start=1):
        cell = ws.cell(row=row, column=j, value=h)
        style_cell(cell, fill=FILL_HEADER, font=FONT_HEADER, align=ALIGN_CENTER)


def ecrire_df_dans_feuille(ws, df: pl.DataFrame, start_row: int = 4) -> None:
    ecrire_entetes(ws, df.columns, start_row)
    for i, row_data in enumerate(df.iter_rows(named=True), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=i, column=j, value=row_data[col])
    ws.freeze_panes = f"A{start_row + 1}"


def ajouter_lien_retour(ws, row: int = 2) -> None:
    cell = ws.cell(row=row, column=1,
                   value='=HYPERLINK("#SOMMAIRE!A1","← Retour au sommaire")')
    style_cell(cell, font=FONT_LINK, align=ALIGN_LEFT)


def ajuster_largeurs(ws, largeurs: dict[int, int]) -> None:
    for col_idx, width in largeurs.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ============================================================
# PRÉPARATION DES DONNÉES PAR SECTION
# ============================================================

print("🔨 Préparation des données par section...")


def extract_grand_groupe(code_citp) -> int | None:
    if code_citp is None:
        return None
    chiffres = re.sub(r"[^0-9]", "", str(code_citp))
    if not chiffres:
        return None
    g = int(chiffres[0])
    return g if 1 <= g <= 9 else None


# --- Section 1 : CITP ---
stats_citp: dict[int, dict] = {}
if dt_citp is not None and "Code_CITP" in dt_citp.columns:
    dt_citp = dt_citp.with_columns(
        pl.col("Code_CITP")
        .map_elements(extract_grand_groupe, return_dtype=pl.Int32)
        .alias("GRAND_GROUPE")
    )
    for gg in range(1, 10):
        sub = dt_citp.filter(pl.col("GRAND_GROUPE") == gg)
        stats_citp[gg] = {
            "effectif_total": int(sub["Effectif"].sum()) if len(sub) > 0 else 0,
            "nb_metiers":     sub["Code_CITP"].n_unique() if len(sub) > 0 else 0,
        }
    print(f"  ✓ CITP : {sum(1 for v in stats_citp.values() if v['effectif_total'] > 0)} grands groupes avec données")

# --- Section 2 : Grade ---
stats_grade: dict[str, dict] = {}
if dt_grade is not None and "GRADE_PRINCIPAL" in dt_grade.columns:
    for grade_val in sorted(dt_grade["GRADE_PRINCIPAL"].unique().to_list()):
        sub = dt_grade.filter(pl.col("GRADE_PRINCIPAL") == grade_val)
        eff_total = int(sub["Effectif"].sum())
        rev_pond = (
            float((sub["Effectif"] * sub["Revenu_moyen"]).sum() / sub["Effectif"].sum())
            if sub["Effectif"].sum() > 0 else 0.0
        )
        stats_grade[grade_val] = {
            "effectif_total":     eff_total,
            "revenu_moyen_ponde": round(rev_pond),
        }
    print(f"  ✓ Grade : {len(stats_grade)} grades")

# --- Section 3 : Sexe ---
stats_sexe: dict[str, dict] = {}
if dt_grade_sexe is not None and "SEXE_STD" in dt_grade_sexe.columns:
    for sexe_val in dt_grade_sexe["SEXE_STD"].unique().to_list():
        sub = dt_grade_sexe.filter(pl.col("SEXE_STD") == sexe_val)
        eff_total = int(sub["Effectif"].sum())
        rev_pond = (
            float((sub["Effectif"] * sub["Revenu_moyen"]).sum() / sub["Effectif"].sum())
            if sub["Effectif"].sum() > 0 else 0.0
        )
        stats_sexe[sexe_val] = {
            "effectif_total":     eff_total,
            "revenu_moyen_ponde": round(rev_pond),
        }
    print(f"  ✓ Sexe : {len(stats_sexe)} catégories")

# --- Section 4 : Multi-postes ---
grades_multi: list[str] = []
if dt_multi is not None and "GRADE_PRINCIPAL" in dt_multi.columns:
    grades_multi = sorted(dt_multi["GRADE_PRINCIPAL"].unique().to_list())
    print(f"  ✓ Multi-postes : {len(grades_multi)} grades")

# --- Section 5 : Barème ---
baremes_liste: list[str] = []
stats_bareme: dict[str, dict] = {}
if BAREME_DISPONIBLE and "bareme" in dt_bareme.columns:
    baremes_liste = sorted(dt_bareme["bareme"].drop_nulls().unique().to_list())
    for b in baremes_liste:
        sub = dt_bareme.filter(pl.col("bareme") == b)
        eff = int(sub["Effectif"].sum()) if "Effectif" in sub.columns else 0
        rev = (
            float((sub["Effectif"] * sub["Revenu_moyen"]).sum() / sub["Effectif"].sum())
            if "Effectif" in sub.columns and sub["Effectif"].sum() > 0 else 0.0
        )
        stats_bareme[b] = {
            "effectif_total":     eff,
            "revenu_moyen_ponde": round(rev),
        }
    print(f"  ✓ Barème : {len(baremes_liste)} grilles salariales")
else:
    print("  ℹ️  Feuilles barème absentes — section 5 ignorée")

print()

# ============================================================
# CRÉATION DU WORKBOOK
# ============================================================

print("📄 Création du workbook...")
wb = Workbook()
wb.remove(wb.active)

# ============================================================
# FEUILLE SOMMAIRE
# ============================================================

print("📋 Création du SOMMAIRE...")
ws_som = wb.create_sheet("SOMMAIRE")

titre_global = (
    f"INDICATEURS SALAIRES — "
    f"{'Salaire brut' if MODE_SALAIRE_BRUT else 'Revenu salarial'} | "
    f"{'2024-2025' if PERIODE_RECENTE else '2015 à récent'} | "
    f"{'avec zéros' if INCLURE_ZEROS else 'sans zéros'}"
)
merge_titre(ws_som, titre_global, row=1, n_cols=4, fill=FILL_TITLE, font=FONT_TITLE)
ws_som.cell(row=2, column=1,
            value="Cliquez sur une section pour accéder aux détails").font = Font(italic=True)
ws_som.merge_cells("A2:D2")

current_row = 4

# ────────────────────────────────────────────────────────────
# SECTION 1 : CITP
# ────────────────────────────────────────────────────────────
if dt_citp is not None:
    merge_titre(ws_som, "📊 SECTION 1 : CLASSIFICATION PAR GRAND GROUPE CITP",
                row=current_row, fill=FILL_SUBTITLE, font=FONT_SUBTITLE)
    current_row += 1
    ecrire_entetes(ws_som, ["Code", "Grand Groupe CITP", "Effectif total", "Nb métiers"],
                   current_row)
    current_row += 1

    for gg in range(1, 10):
        titre_gg   = CITP_GRANDS_GROUPES[gg]
        eff        = stats_citp.get(gg, {}).get("effectif_total", 0)
        nb_m       = stats_citp.get(gg, {}).get("nb_metiers", 0)
        sheet_name = f"CITP_GG{gg}"

        ws_som.cell(row=current_row, column=1, value=gg)
        lien_interne(ws_som, current_row, 2, titre_gg, sheet_name)
        ws_som.cell(row=current_row, column=3, value=eff)
        ws_som.cell(row=current_row, column=4, value=nb_m)
        current_row += 1

    current_row += 2

# ────────────────────────────────────────────────────────────
# SECTION 2 : GRADE
# ────────────────────────────────────────────────────────────
if dt_grade is not None:
    merge_titre(ws_som, "📊 SECTION 2 : CLASSIFICATION PAR GRADE",
                row=current_row, fill=FILL_SUBTITLE, font=FONT_SUBTITLE)
    current_row += 1
    ecrire_entetes(ws_som, ["Grade", "Description", "Effectif total", "Revenu moyen"],
                   current_row)
    current_row += 1

    for grade_val in sorted(stats_grade.keys()):
        cat = (
            "Catégorie A" if grade_val.startswith("A") else
            "Catégorie B" if grade_val.startswith("B") else
            "Catégorie C" if grade_val.startswith("C") else
            "Catégorie D" if grade_val.startswith("D") else "Autre"
        )
        sheet_name = f"GRADE_{grade_val}"
        lien_interne(ws_som, current_row, 1, grade_val, sheet_name)
        ws_som.cell(row=current_row, column=2, value=cat)
        ws_som.cell(row=current_row, column=3, value=stats_grade[grade_val]["effectif_total"])
        ws_som.cell(row=current_row, column=4, value=stats_grade[grade_val]["revenu_moyen_ponde"])
        current_row += 1

    current_row += 2

# ────────────────────────────────────────────────────────────
# SECTION 3 : SEXE
# ────────────────────────────────────────────────────────────
if dt_grade_sexe is not None:
    merge_titre(ws_som, "📊 SECTION 3 : CLASSIFICATION PAR SEXE",
                row=current_row, fill=FILL_SUBTITLE, font=FONT_SUBTITLE)
    current_row += 1
    ecrire_entetes(ws_som, ["Sexe", "Description", "Effectif total", "Revenu moyen"],
                   current_row)
    current_row += 1

    for sexe_val in sorted(stats_sexe.keys()):
        desc = (
            "Agents masculins" if sexe_val == "Homme" else
            "Agents féminins"  if sexe_val == "Femme" else
            "Non renseigné"
        )
        sheet_name = f"SEXE_{sexe_val}"
        lien_interne(ws_som, current_row, 1, sexe_val, sheet_name)
        ws_som.cell(row=current_row, column=2, value=desc)
        ws_som.cell(row=current_row, column=3, value=stats_sexe[sexe_val]["effectif_total"])
        ws_som.cell(row=current_row, column=4, value=stats_sexe[sexe_val]["revenu_moyen_ponde"])
        current_row += 1

    current_row += 2

# ────────────────────────────────────────────────────────────
# SECTION 4 : MULTI-POSTES
# ────────────────────────────────────────────────────────────
if dt_multi is not None:
    merge_titre(ws_som, "📊 SECTION 4 : MULTI-POSTES PAR GRADE",
                row=current_row, fill=FILL_SUBTITLE, font=FONT_SUBTITLE)
    current_row += 1
    ecrire_entetes(ws_som, ["Grade", "Description", "Mono-postes", "Multi-postes"],
                   current_row)
    current_row += 1

    for grade_val in grades_multi:
        sub   = dt_multi.filter(pl.col("GRADE_PRINCIPAL") == grade_val)
        mono  = int(sub["1"].sum())  if "1"  in sub.columns else 0
        multi = sum(
            int(sub[c].sum()) if c in sub.columns else 0
            for c in ["2", "3", "4", "5+"]
        )
        cat = (
            "Catégorie A" if grade_val.startswith("A") else
            "Catégorie B" if grade_val.startswith("B") else
            "Catégorie C" if grade_val.startswith("C") else
            "Catégorie D" if grade_val.startswith("D") else "Autre"
        )
        sheet_name = f"MULTI_{grade_val}"
        lien_interne(ws_som, current_row, 1, grade_val, sheet_name)
        ws_som.cell(row=current_row, column=2, value=cat)
        ws_som.cell(row=current_row, column=3, value=mono)
        ws_som.cell(row=current_row, column=4, value=multi)
        current_row += 1

# ────────────────────────────────────────────────────────────
# SECTION 5 : BARÈME / GRILLE SALARIALE
# ────────────────────────────────────────────────────────────
if BAREME_DISPONIBLE:
    merge_titre(ws_som, "📊 SECTION 5 : CLASSIFICATION PAR BARÈME / GRILLE SALARIALE",
                row=current_row, fill=FILL_SUBTITLE, font=FONT_SUBTITLE)
    current_row += 1
    ecrire_entetes(ws_som, ["Barème / Grille salariale", "Effectif total", "Revenu moyen",
                             "Détail grade", "Détail sexe"],
                   current_row)
    current_row += 1

    for b in baremes_liste:
        sheet_det   = f"BAREME_{re.sub(r'[^A-Za-z0-9]', '_', b)[:25]}"
        sheet_grade = f"BAREME_G_{re.sub(r'[^A-Za-z0-9]', '_', b)[:20]}"
        sheet_sexe  = f"BAREME_S_{re.sub(r'[^A-Za-z0-9]', '_', b)[:20]}"
        lien_interne(ws_som, current_row, 1, b, sheet_det)
        ws_som.cell(row=current_row, column=2,
                    value=stats_bareme.get(b, {}).get("effectif_total", ""))
        ws_som.cell(row=current_row, column=3,
                    value=stats_bareme.get(b, {}).get("revenu_moyen_ponde", ""))
        lien_interne(ws_som, current_row, 4, "→ par grade", sheet_grade)
        lien_interne(ws_som, current_row, 5, "→ par sexe",  sheet_sexe)
        current_row += 1

    current_row += 1
    # Lien MODELE_ANSTAT
    if dt_modele_anstat is not None:
        merge_titre(ws_som, "📋 MODÈLE STATISTIQUE ANSTAT — Format hiérarchique",
                    row=current_row, n_cols=5, fill=PatternFill("solid", fgColor="1F4E79"),
                    font=Font(color="FFFFFF", bold=True, size=11))
        current_row += 1
        lien_interne(ws_som, current_row, 1, "Voir le modèle ANSTAT →", "MODELE_ANSTAT")
        current_row += 2

# Formatage colonnes sommaire
ajuster_largeurs(ws_som, {1: 45, 2: 18, 3: 18, 4: 18, 5: 18})
print("  ✓ SOMMAIRE créé\n")

# ============================================================
# FEUILLES DE DÉTAIL
# ============================================================

print("📄 Création des feuilles détail...")

# ── Section 1 : CITP par grand groupe ───────────────────────
if dt_citp is not None:
    for gg in range(1, 10):
        sheet_name = f"CITP_GG{gg}"
        gg_data    = dt_citp.filter(pl.col("GRAND_GROUPE") == gg).drop("GRAND_GROUPE")
        if len(gg_data) == 0:
            continue
        ws = wb.create_sheet(sheet_name)
        merge_titre(ws, CITP_GRANDS_GROUPES[gg], row=1, n_cols=max(len(gg_data.columns), 4))
        ajouter_lien_retour(ws, row=2)
        ecrire_df_dans_feuille(ws, gg_data, start_row=4)
        ajuster_largeurs(ws, {1: 20, 2: 40, 3: 15, 4: 15, 5: 15, 6: 15, 7: 15, 8: 15})
        print(f"  ✓ {sheet_name}")

# ── Section 2 : Grade ────────────────────────────────────────
if dt_grade is not None:
    for grade_val in sorted(dt_grade["GRADE_PRINCIPAL"].unique().to_list()):
        sheet_name = f"GRADE_{grade_val}"
        data       = dt_grade.filter(pl.col("GRADE_PRINCIPAL") == grade_val)
        ws = wb.create_sheet(sheet_name)
        merge_titre(ws, f"GRADE {grade_val}", row=1, n_cols=max(len(data.columns), 4))
        ajouter_lien_retour(ws, row=2)
        ecrire_df_dans_feuille(ws, data, start_row=4)
        ajuster_largeurs(ws, {1: 15, 2: 15, 3: 15, 4: 18, 5: 18, 6: 15, 7: 15, 8: 15})
        print(f"  ✓ {sheet_name}")

# ── Section 3 : Sexe ────────────────────────────────────────
if dt_grade_sexe is not None:
    for sexe_val in dt_grade_sexe["SEXE_STD"].unique().to_list():
        sheet_name = f"SEXE_{sexe_val}"
        data       = dt_grade_sexe.filter(pl.col("SEXE_STD") == sexe_val)
        ws = wb.create_sheet(sheet_name)
        merge_titre(ws, f"SEXE : {sexe_val}", row=1, n_cols=max(len(data.columns), 4))
        ajouter_lien_retour(ws, row=2)
        ecrire_df_dans_feuille(ws, data, start_row=4)
        ajuster_largeurs(ws, {1: 15, 2: 15, 3: 15, 4: 18, 5: 18, 6: 15, 7: 15, 8: 15})
        print(f"  ✓ {sheet_name}")

# ── Section 4 : Multi-postes par grade ──────────────────────
if dt_multi is not None:
    for grade_val in grades_multi:
        sheet_name = f"MULTI_{grade_val}"
        data       = dt_multi.filter(pl.col("GRADE_PRINCIPAL") == grade_val)
        ws = wb.create_sheet(sheet_name)
        merge_titre(ws, f"MULTI-POSTES : {grade_val}", row=1, n_cols=max(len(data.columns), 4))
        ajouter_lien_retour(ws, row=2)
        ecrire_df_dans_feuille(ws, data, start_row=4)
        ajuster_largeurs(ws, {1: 15, 2: 15, 3: 15, 4: 15, 5: 12, 6: 12, 7: 12, 8: 12})
        print(f"  ✓ {sheet_name}")

# ── Section 5 : Barème ──────────────────────────────────────
if BAREME_DISPONIBLE:
    # Détail par barème (revenu)
    for b in baremes_liste:
        sheet_name = f"BAREME_{re.sub(r'[^A-Za-z0-9]', '_', b)[:25]}"
        data = dt_bareme.filter(pl.col("bareme") == b)
        if len(data) == 0:
            continue
        ws = wb.create_sheet(sheet_name)
        merge_titre(ws, f"BARÈME : {b}", row=1, n_cols=max(len(data.columns), 5))
        ajouter_lien_retour(ws, row=2)
        ecrire_df_dans_feuille(ws, data, start_row=4)
        ajuster_largeurs(ws, {1: 20, 2: 15, 3: 15, 4: 18, 5: 18, 6: 15, 7: 15, 8: 15})
        print(f"  ✓ {sheet_name}")

    # Détail par barème × grade
    if dt_bareme_grade is not None:
        for b in baremes_liste:
            sheet_name = f"BAREME_G_{re.sub(r'[^A-Za-z0-9]', '_', b)[:20]}"
            data = dt_bareme_grade.filter(pl.col("bareme") == b)
            if len(data) == 0:
                continue
            ws = wb.create_sheet(sheet_name)
            merge_titre(ws, f"BARÈME × GRADE : {b}", row=1, n_cols=max(len(data.columns), 5))
            ajouter_lien_retour(ws, row=2)
            ecrire_df_dans_feuille(ws, data, start_row=4)
            ajuster_largeurs(ws, {1: 20, 2: 15, 3: 15, 4: 15, 5: 18, 6: 18, 7: 15, 8: 15})
            print(f"  ✓ {sheet_name}")

    # Détail par barème × sexe
    if dt_bareme_sexe is not None:
        for b in baremes_liste:
            sheet_name = f"BAREME_S_{re.sub(r'[^A-Za-z0-9]', '_', b)[:20]}"
            data = dt_bareme_sexe.filter(pl.col("bareme") == b)
            if len(data) == 0:
                continue
            ws = wb.create_sheet(sheet_name)
            merge_titre(ws, f"BARÈME × SEXE : {b}", row=1, n_cols=max(len(data.columns), 5))
            ajouter_lien_retour(ws, row=2)
            ecrire_df_dans_feuille(ws, data, start_row=4)
            ajuster_largeurs(ws, {1: 20, 2: 15, 3: 15, 4: 15, 5: 18, 6: 18, 7: 15, 8: 15})
            print(f"  ✓ {sheet_name}")

    # MODELE_ANSTAT : copie directe depuis la feuille source
    if dt_modele_anstat is not None and len(dt_modele_anstat) > 0:
        ws = wb.create_sheet("MODELE_ANSTAT")
        merge_titre(ws, "MODÈLE DE DONNÉES DE SALAIRE DES FONCTIONNAIRES",
                    row=1, n_cols=max(len(dt_modele_anstat.columns), 4),
                    fill=PatternFill("solid", fgColor="1F4E79"),
                    font=Font(color="FFFFFF", bold=True, size=12))
        ajouter_lien_retour(ws, row=2)
        ecrire_df_dans_feuille(ws, dt_modele_anstat, start_row=4)
        ajuster_largeurs(ws, {1: 55, 2: 22, 3: 22, 4: 22})
        print(f"  ✓ MODELE_ANSTAT")

print()

# ============================================================
# SAUVEGARDE ET UPLOAD
# ============================================================

print("💾 Sauvegarde...")

buf_out    = io.BytesIO()
wb.save(buf_out)
buf_out.seek(0)
xlsx_bytes = buf_out.getvalue()

s3.put_object(
    Bucket=BUCKET_STAGING,
    Key=OUTPUT_KEY,
    Body=xlsx_bytes,
    ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

taille_mb = len(xlsx_bytes) / 1024**2

print(f"\n✅ Fichier créé : s3://{BUCKET_STAGING}/{OUTPUT_KEY}")
print(f"   Taille   : {taille_mb:.1f} MB")
print(f"   Feuilles : {len(wb.sheetnames)}\n")

print("=" * 70)
print("✅ TERMINÉ")
print("=" * 70 + "\n")
print("📊 STRUCTURE DU FICHIER :\n")
print("FEUILLE SOMMAIRE (4 sections cliquables) :")
print("  • Section 1 : Classification par grand groupe CITP (9 groupes)")
print("  • Section 2 : Classification par grade (A1-D3)")
print("  • Section 3 : Classification par sexe (Homme/Femme)")
print("  • Section 4 : Multi-postes par grade\n")
print("💡 UTILISATION :")
print("  1. Ouvrez le fichier Excel")
print("  2. La feuille SOMMAIRE s'affiche avec 4 sections")
print("  3. Cliquez sur n'importe quel lien bleu pour naviguer")
print("  4. Utilisez '← Retour au sommaire' pour revenir\n")