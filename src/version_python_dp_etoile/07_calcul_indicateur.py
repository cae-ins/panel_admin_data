# ============================================================
# PANEL ADMIN — ÉTAPE 7 : CALCUL DES INDICATEURS SALAIRES
# ============================================================
# Produit les indicateurs statistiques complets sur les salaires
# à partir des fichiers Gold panel.
#
# Prérequis : 06b_imputer_grades.py doit avoir été exécuté.
# Les colonnes GRADE et GRADE_SOURCE sont lues directement
# depuis le Gold — aucune imputation n'est recalculée ici.
#
# Paramètres configurables :
#   MODE_SALAIRE_BRUT : True  = agrégation par cle_unique (salaire brut)
#                       False = agrégation par matricule (revenu salarial)
#   PERIODE_RECENTE   : True  = années récentes uniquement (2024-2025)
#                       False = période complète (2015 → dernière année)
#   INCLURE_ZEROS     : True  = inclure salaires = 0 ou NA
#                       False = salaire strictement > 0
#
# Tables produites (Excel multi-feuilles) :
#   REVENU_Grade_Detail      : Effectif, moy, méd, min, max, P25, P75 par grade
#   REVENU_Grade_Sexe        : Idem, par grade × sexe
#   REVENU_CITP_Detail       : Idem, par CITP
#   SALAIRE_BRUT_Moyen       : Salaire brut moyen par année × mois × grade (pivot)
#   SALAIRE_BRUT_Median      : Salaire brut médian par année × mois × grade (pivot)
#   MULTI_Par_Grade_NbPostes : Effectifs par grade × nombre de postes (pivot)
#   MULTI_Distribution       : Distribution globale mono/multi-postes
#
# Source  : s3://gold/panel_admin/panel_YYYY.parquet
# Sortie  : s3://staging/panel_admin/exports_gold/indicateurs_*.xlsx
#
# Dépendances :
#   pip install polars boto3 python-dotenv openpyxl
# ============================================================

import gc
import io
import os
import re
import tempfile
import unicodedata
from dotenv import load_dotenv
import boto3
from botocore.client import Config
import polars as pl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

load_dotenv(".env")

# ============================================================
# PARAMÈTRES DE CONFIGURATION
# ============================================================

# True = Salaire brut (cle_unique) | False = Revenu salarial (matricule)
MODE_SALAIRE_BRUT = True

# True = 2024-2025 | False = 2015 → dernière année disponible
PERIODE_RECENTE   = True

# True = inclure zéros | False = salaire strictement > 0
INCLURE_ZEROS     = True

# Seuil de winsorisation (ex: 0.01 = écrêtage à P1/P99)
# Mettre à None pour désactiver
SEUIL_WINS        = 0.01

# Effectif minimum par cellule mois × grade × CITP pour appliquer la
# winsorisation CITP. En dessous, repli sur mois × grade uniquement.
MIN_OBS_WINS      = 100

# ============================================================

# LORSQU'ON TRAVAILLE DEPUIS SA MACHINE LOCAL
MINIO_ENDPOINT   = os.getenv("MINIO_ENDPOINT",   "http://192.168.1.230:30137")
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY", "datalab-team")
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY", "minio-datalabteam123")

BUCKET_GOLD    = os.getenv("BUCKET_GOLD",    "gold")
BUCKET_STAGING = os.getenv("BUCKET_STAGING", "staging")
PREFIX_PANEL   = "panel_admin"
PREFIX_EXPORTS = "panel_admin/exports_gold"
KEY_CORRESPONDANCE = "panel_admin/references/CORRESPONDANCE_GRILLE_EMPLOI.xlsx"

_mode_label    = "salaire_brut"  if MODE_SALAIRE_BRUT else "revenu_salarial"
_periode_label = "2024_2025"     if PERIODE_RECENTE   else "2015_recent"
_zeros_label   = "avec_zeros"    if INCLURE_ZEROS     else "sans_zeros"
OUTPUT_KEY     = (f"{PREFIX_EXPORTS}/indicateurs_{_mode_label}"
                  f"_{_periode_label}_{_zeros_label}.xlsx")

# Libellés mois français
MOIS_FR = {
    1: "Janvier",  2: "Février",   3: "Mars",      4: "Avril",
    5: "Mai",      6: "Juin",      7: "Juillet",   8: "Août",
    9: "Septembre",10: "Octobre",  11: "Novembre", 12: "Décembre",
}

GRADES_VALIDES = (
    [f"A{i}" for i in range(1, 8)] +
    [f"B{i}" for i in range(1, 7)] +
    [f"C{i}" for i in range(1, 6)] +
    [f"D{i}" for i in range(1, 4)]
)

# Composantes du salaire brut (codes F2, suffixe _f2 dans le Gold)
COMPOSANTES_SALAIRE = ["124_f2", "228_f2", "271_f2", "421_f2", "425_f2"]

# --- CLIENT S3 ---
s3 = boto3.client(
    "s3",
    endpoint_url          = MINIO_ENDPOINT,
    aws_access_key_id     = MINIO_ACCESS_KEY,
    aws_secret_access_key = MINIO_SECRET_KEY,
    config                = Config(
        signature_version = "s3v4",
        retries           = {"max_attempts": 5, "mode": "adaptive"},
        connect_timeout   = 30,
        read_timeout      = 120,
    ),
    region_name = "us-east-1",
    verify      = False,
)

print("=" * 70)
print(f"INDICATEURS SALAIRES — Mode : "
      f"{'Salaire brut (cle_unique)' if MODE_SALAIRE_BRUT else 'Revenu salarial (matricule)'}"
      f" | Période : {'2024-2025' if PERIODE_RECENTE else '2015 → récent'}")
print("=" * 70)
print()

# ============================================================
# FONCTIONS UTILITAIRES
# ============================================================

def lire_parquet_s3(bucket: str, key: str) -> pl.DataFrame:
    buf = io.BytesIO()
    s3.download_fileobj(bucket, key, buf)
    buf.seek(0)
    return pl.read_parquet(buf)


# ============================================================
# CONSTANTES BARÈME
# ============================================================

# Pour ENS SEC & PRIM : grades B/C/D = Primaire, grades A = Secondaire
GRADES_PRIMAIRE = {"B3", "C3", "D1", "D2"}

# Ordre d'affichage dans la feuille MODELE_ANSTAT
ORDRE_MODELE_ANSTAT = [
    ("TOUS SECTEURS CONFONDUS",                       "_GLOBAL_"),
    ("SECTEUR EDUCATION",                             None),
    ("Primaire",                                      "ENS SEC & PRIM_PRIM"),
    ("Secondaire",                                    "ENS SEC & PRIM_SEC"),
    ("Supérieur",                                     "ENS SUP"),
    ("SECTEUR SANTE",                                 None),
    ("Cadres supérieurs",                             "Cadre-Sup-Santé Generaliste & Specialiste"),
    ("Personnels Techniques",                         "Tech-Santé"),
    ("AUTRES SECTEURS",                               None),
    ("Para-militaire / Douanes / Eaux & Forêts",      "police"),
    ("Barème Général",                                "Barême Général"),
    ("PERSONNELS SOUS STATUT PARTICULIER",            None),
    ("Magistrat",                                     "MAGISTRAT "),
    ("Corps préfectoral",                             "CORPS PREF"),
    ("Corps diplomatique",                            "CORPS DIPL"),
    ("Greffier",                                      "GREFFE"),
]


def normaliser_pour_matching(texte) -> str | None:
    """Normalisation textuelle pour fuzzy join — identique à l'étape 03."""
    if texte is None:
        return None
    texte = str(texte).strip()
    if not texte:
        return None
    texte = re.sub(r"[-]", " ", texte)
    texte = re.sub(r"[.,;:/()\\[\]]", " ", texte)
    texte = re.sub(r"[''\u2019]", " ", texte)
    texte = texte.upper()
    texte = unicodedata.normalize("NFKD", texte)
    texte = "".join(c for c in texte if not unicodedata.combining(c))
    texte = re.sub(r"\s+", " ", texte).strip()
    return texte if texte else None


def charger_table_bareme() -> pl.DataFrame | None:
    """
    Charge la table CORRESPONDANCE_GRILLE_EMPLOI depuis staging et
    renvoie une table de jointure (emploi_norm → bareme, grade_ref).
    Renvoie None si le fichier est absent (étape non bloquante).
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        s3.download_file(BUCKET_STAGING, KEY_CORRESPONDANCE, tmp_path)
        df = pl.read_excel(tmp_path)
        df = df.rename({
            "BAREME / GRILLE SALARIALE": "bareme",
            "GRADE": "grade_ref",
            "EMPLOI": "emploi_ref",
        })
        df = df.with_columns(
            pl.col("emploi_ref")
            .map_elements(normaliser_pour_matching, return_dtype=pl.Utf8)
            .alias("emploi_ref_norm")
        )
        table = (
            df.filter(pl.col("emploi_ref_norm").is_not_null())
            .select(["emploi_ref_norm", "bareme", "grade_ref"])
            .unique(subset=["emploi_ref_norm"])
        )
        print(f"  ✓ Table barème chargée : {len(table)} emplois")
        return table
    except Exception as e:
        print(f"  ⚠️  Table barème non disponible ({e}) — feuilles barème ignorées")
        return None
    finally:
        os.unlink(tmp_path)


def fuzzy_join_bareme(df: pl.DataFrame, table_join: pl.DataFrame,
                      col_emploi: str) -> pl.DataFrame:
    """
    Ajoute les colonnes 'bareme' et 'grade_ref' au DataFrame
    via fuzzy join sur le libellé emploi normalisé.
    """
    df = df.with_columns(
        pl.col(col_emploi)
        .map_elements(normaliser_pour_matching, return_dtype=pl.Utf8)
        .alias("_emploi_norm")
    )
    df = df.join(table_join, left_on="_emploi_norm", right_on="emploi_ref_norm",
                 how="left").drop("_emploi_norm")
    nb_match  = int(df["bareme"].is_not_null().sum())
    taux      = 100 * nb_match / len(df)
    print(f"  ✓ Fuzzy join barème : {nb_match:,} / {len(df):,} ({taux:.1f}%)")
    if taux < 20:
        print("  ⚠️  Matching faible — vérifiez CORRESPONDANCE_GRILLE_EMPLOI.xlsx")
    return df


def ajouter_bareme_detail(df: pl.DataFrame, col_grade: str) -> pl.DataFrame:
    """Scinde ENS SEC & PRIM en Primaire (B/C/D) et Secondaire (A)."""
    return df.with_columns(
        pl.when(
            (pl.col("bareme") == "ENS SEC & PRIM")
            & pl.col(col_grade).is_in(GRADES_PRIMAIRE)
        ).then(pl.lit("ENS SEC & PRIM_PRIM"))
        .when(
            (pl.col("bareme") == "ENS SEC & PRIM")
            & ~pl.col(col_grade).is_in(GRADES_PRIMAIRE)
        ).then(pl.lit("ENS SEC & PRIM_SEC"))
        .otherwise(pl.col("bareme"))
        .alias("bareme_detail")
    )


def grade_order_value(grade: str) -> int:
    """Valeur ordinale d'un grade pour sélectionner le grade principal."""
    if re.match(r"^[ABCD][0-9]$", grade):
        base = {"A": 80, "B": 70, "C": 60, "D": 50}[grade[0]]
        return base + int(grade[1])
    return 0


def compute_indicators(df: pl.DataFrame, group_cols: list[str],
                       value_col: str = "REVENU_TOTAL") -> pl.DataFrame:
    """Calcule les indicateurs statistiques complets par groupe."""
    return (
        df.group_by(group_cols)
        .agg([
            pl.len().alias("Effectif"),
            pl.col(value_col).mean().round(0).alias("Revenu_moyen"),
            pl.col(value_col).median().round(0).alias("Revenu_median"),
            pl.col(value_col).min().round(0).alias("Revenu_min"),
            pl.col(value_col).max().round(0).alias("Revenu_max"),
            pl.col(value_col).quantile(0.25, interpolation="nearest")
              .round(0).alias("Revenu_p25"),
            pl.col(value_col).quantile(0.75, interpolation="nearest")
              .round(0).alias("Revenu_p75"),
        ])
        .sort(group_cols)
    )


# ============================================================
# 1. CHARGEMENT
# ============================================================

print("1. Chargement des fichiers Gold panel...")

pages_panel = s3.get_paginator("list_objects_v2").paginate(
    Bucket=BUCKET_GOLD, Prefix=PREFIX_PANEL
)
cles_panel = sorted([
    obj["Key"]
    for page in pages_panel
    for obj in page.get("Contents", [])
    if re.search(r"panel_20\d{2}\.parquet$", obj["Key"])
])

print(f"   Fichiers panel trouvés : {len(cles_panel)}")

frames = []
for cle in cles_panel:
    print(f"   Lecture {os.path.basename(cle)}...", end=" ", flush=True)
    df_part = lire_parquet_s3(BUCKET_GOLD, cle)
    frames.append(df_part)
    print(f"✓ {len(df_part):,} lignes")

base = pl.concat(frames, how="diagonal")
del frames
gc.collect()
print(f"   ✓ Total : {len(base):,} lignes × {len(base.columns)} colonnes\n")

# ============================================================
# 2. VÉRIFICATION GRADE ET GRADE_SOURCE
# ============================================================

print("2. Vérification des colonnes GRADE et GRADE_SOURCE...")

if "GRADE" not in base.columns or "GRADE_SOURCE" not in base.columns:
    raise RuntimeError(
        "Colonnes GRADE et/ou GRADE_SOURCE absentes du Gold.\n"
        "Exécutez d'abord 06b_imputer_grades.py avant ce script."
    )

n_total = len(base)
for source in ["OBSERVE", "IMPUTE", "NF_NON_IMPUTABLE"]:
    n = int((base["GRADE_SOURCE"] == source).sum())
    print(f"   {source:<20} : {n:,}  ({100*n/n_total:.1f}%)")
print()

# ============================================================
# 3. EXTRACTION ANNÉE / MOIS depuis mois_annee (format MMYYYY)
# ============================================================

print("3. Extraction année/mois depuis mois_annee...")

# mois_annee est au format MMYYYY (ex. "012015")
# mois  = slice(0, 2)  →  "01"
# annee = slice(2, 4)  →  "2015"
base = base.with_columns([
    pl.col("mois_annee").str.slice(2, 4).cast(pl.Int32, strict=False).alias("ANNEE"),
    pl.col("mois_annee").str.slice(0, 2).cast(pl.Int32, strict=False).alias("MOIS_NUM"),
])

annees_dispo = sorted(base["ANNEE"].drop_nulls().unique().to_list())
print(f"   Années disponibles : {annees_dispo}")

ANNEES_ANALYSE = [2024, 2025] if PERIODE_RECENTE else [a for a in annees_dispo if a >= 2015]
print(f"   Période retenue    : {ANNEES_ANALYSE}")

base = base.filter(pl.col("ANNEE").is_in(ANNEES_ANALYSE))
print(f"   Filtrage           : {len(base):,} lignes\n")

if len(base) == 0:
    raise ValueError(f"Aucune donnée pour les années {ANNEES_ANALYSE}")

base = base.with_columns(
    pl.col("MOIS_NUM")
    .map_elements(lambda m: MOIS_FR.get(m, str(m) if m is not None else ""), return_dtype=pl.Utf8)
    .alias("MOIS")
)

# ============================================================
# 4. CALCUL DU SALAIRE BRUT (composantes élémentaires)
# ============================================================

print("4. Calcul du salaire brut depuis les composantes...")

composantes_presentes = [c for c in COMPOSANTES_SALAIRE if c in base.columns]
print(f"   Composantes trouvées : {composantes_presentes}")

if composantes_presentes:
    base = base.with_columns([
        pl.col(c).cast(pl.Float64, strict=False) for c in composantes_presentes
    ])
    base = base.with_columns(
        pl.sum_horizontal([pl.col(c) for c in composantes_presentes])
        .alias("SALAIRE_BRUT")
    )
else:
    base = base.with_columns(
        pl.col("montant_brut").cast(pl.Float64, strict=False).alias("SALAIRE_BRUT")
        if "montant_brut" in base.columns
        else pl.lit(0.0).alias("SALAIRE_BRUT")
    )

n_pos = int((base["SALAIRE_BRUT"] > 0).sum())
print(f"   Salaires > 0 : {n_pos:,} ({100*n_pos/len(base):.1f}%)\n")

# ============================================================
# 5. STANDARDISATION DU SEXE
# ============================================================

print("5. Standardisation sexe...")

if "sexe" in base.columns:
    base = base.with_columns(
        pl.col("sexe").cast(pl.Utf8)
        .str.to_uppercase()
        .str.strip_chars()
        .map_elements(
            lambda s: (
                "Homme" if s in {"MASCULIN", "M", "H", "HOMME", "MALE", "1"} else
                "Femme" if s in {"FEMININ", "F", "FEMME", "FEMALE", "2",
                                 "FÉMININ", "FEMININE"} else
                "Inconnu"
            ),
            return_dtype=pl.Utf8
        )
        .alias("SEXE_STD")
    )
    n_inconnu = int((base["SEXE_STD"] == "Inconnu").sum())
    if n_inconnu > 0:
        base = base.filter(pl.col("SEXE_STD") != "Inconnu")
        print(f"   Exclusion sexe inconnu : {n_inconnu:,}")
    for sexe_val in ["Homme", "Femme"]:
        n = int((base["SEXE_STD"] == sexe_val).sum())
        print(f"   {sexe_val} : {n:,}")
else:
    base = base.with_columns(pl.lit("Non renseigné").alias("SEXE_STD"))
print()

# ============================================================
# 6. DÉTECTION DES MULTI-POSTES
# ============================================================

print("6. Détection multi-postes (par matricule)...")

if "matricule" in base.columns:
    nb_postes = (
        base.group_by(["mois_annee", "matricule"])
        .agg(pl.len().alias("NB_POSTES"))
    )
    base = base.join(nb_postes, on=["mois_annee", "matricule"], how="left")
    total_multi  = int(
        base.filter(pl.col("NB_POSTES") > 1)
        .select(["mois_annee", "matricule"]).unique().select(pl.len()).item()
    )
    total_agents = int(
        base.select(["mois_annee", "matricule"]).unique().select(pl.len()).item()
    )
    print(f"   Multi-postes : {total_multi:,} / {total_agents:,} "
          f"({100*total_multi/total_agents:.1f}%)\n")
else:
    base = base.with_columns(pl.lit(1).alias("NB_POSTES"))
    print("   Colonne matricule absente, NB_POSTES = 1\n")

# ============================================================
# 7. FILTRAGE SUR GRADES VALIDES
# ============================================================
# On filtre sur GRADE valide (issu de 06b, OBSERVE ou IMPUTE).
# Les NF_NON_IMPUTABLE sont exclus des calculs d'indicateurs.

print("7. Filtrage grades valides avant agrégation...")

if INCLURE_ZEROS:
    base_gv = base.filter(pl.col("GRADE").is_in(GRADES_VALIDES))
else:
    base_gv = base.filter(
        pl.col("GRADE").is_in(GRADES_VALIDES) & (pl.col("SALAIRE_BRUT") > 0)
    )

n_gv = len(base_gv)
for source in ["OBSERVE", "IMPUTE"]:
    n = int((base_gv["GRADE_SOURCE"] == source).sum())
    print(f"   {source:<10} : {n:,}  ({100*n/n_gv:.1f}%)")
print(f"   Total retenu : {n_gv:,} ({100*n_gv/len(base):.1f}%)\n")

# ============================================================
# 7b. FUZZY JOIN EMPLOI → BARÈME
# ============================================================
# Ajoute les colonnes 'bareme' et 'bareme_detail' sur base_gv
# pour permettre la ventilation des indicateurs par grille salariale.

print("7b. Fuzzy join EMPLOI → BARÈME (via table de correspondance)...")

TABLE_BAREME = charger_table_bareme()

CANDIDATES_EMPLOI = ["emploi", "CODE_EMPLOI", "EMPLOI", "corps", "CORPS"]
col_emploi_panel = next(
    (c for c in CANDIDATES_EMPLOI if c in base_gv.columns), None
)

if TABLE_BAREME is not None and col_emploi_panel is not None:
    base_gv = fuzzy_join_bareme(base_gv, TABLE_BAREME, col_emploi_panel)
    # Utiliser grade_ref si GRADE absent ou pour affiner la scission Primaire/Secondaire
    col_grade_bareme = "GRADE" if "GRADE" in base_gv.columns else "grade_ref"
    if col_grade_bareme in base_gv.columns:
        base_gv = ajouter_bareme_detail(base_gv, col_grade_bareme)
    else:
        base_gv = base_gv.with_columns(pl.col("bareme").alias("bareme_detail"))
    BAREME_DISPONIBLE = True
else:
    if TABLE_BAREME is None:
        print("  ⚠️  Table barème absente — feuilles barème non produites")
    else:
        print(f"  ⚠️  Colonne emploi absente dans le panel — feuilles barème non produites")
        print(f"       Colonnes disponibles : {base_gv.columns}")
    BAREME_DISPONIBLE = False

print()

# ============================================================
# 8. AGRÉGATION → REVENU TOTAL
# ============================================================

agg_col   = "cle_unique" if MODE_SALAIRE_BRUT else "matricule"
agg_label = agg_col
print(f"8. Agrégation par {agg_label} → REVENU_TOTAL...")

if agg_col not in base_gv.columns:
    agg_col = "matricule"

base_gv = base_gv.with_columns(
    pl.col("GRADE").map_elements(grade_order_value, return_dtype=pl.Int32)
    .alias("GRADE_ORDER")
)

cols_agg = ["ANNEE", "MOIS_NUM", "MOIS", agg_col]

base_agrege = (
    base_gv
    .group_by(cols_agg)
    .agg([
        pl.col("SALAIRE_BRUT").sum().alias("REVENU_TOTAL"),
        pl.col("NB_POSTES").first().alias("NB_POSTES"),
        pl.col("GRADE").sort_by("GRADE_ORDER", descending=True).first().alias("GRADE_PRINCIPAL"),
        pl.col("GRADE_SOURCE").sort_by("GRADE_ORDER", descending=True).first().alias("GRADE_SOURCE_PRINCIPAL"),
        pl.col("SEXE_STD").first().alias("SEXE_STD"),
        pl.col("Code_CITP").first().alias("Code_CITP")
          if "Code_CITP" in base_gv.columns
          else pl.lit(None).cast(pl.Utf8).alias("Code_CITP"),
        pl.col("Metier_CITP").first().alias("Metier_CITP")
          if "Metier_CITP" in base_gv.columns
          else pl.lit(None).cast(pl.Utf8).alias("Metier_CITP"),
        # Barème : conservé directement depuis base_gv (ajouté en 7b)
        pl.col("bareme").first().alias("bareme")
          if "bareme" in base_gv.columns
          else pl.lit(None).cast(pl.Utf8).alias("bareme"),
        pl.col("bareme_detail").first().alias("bareme_detail")
          if "bareme_detail" in base_gv.columns
          else pl.lit(None).cast(pl.Utf8).alias("bareme_detail"),
    ])
)

base_gv = base_gv.drop("GRADE_ORDER")

if not INCLURE_ZEROS:
    base_agrege = base_agrege.filter(pl.col("REVENU_TOTAL") > 0)
base_agrege = base_agrege.filter(pl.col("REVENU_TOTAL").is_not_null())

print(f"   {len(base_agrege):,} agents agrégés\n")

# ============================================================
# 8b. WINSORISATION DU REVENU TOTAL
# ============================================================
# Appliquée uniquement en mémoire pour le calcul des indicateurs.
# Le Gold n'est pas modifié — REVENU_TOTAL brut est conservé.

value_col_indic = "REVENU_TOTAL"

if SEUIL_WINS is not None:
    wins_full     = [c for c in ["ANNEE", "MOIS_NUM", "GRADE_PRINCIPAL", "Code_CITP"]
                     if c in base_agrege.columns]
    wins_fallback = [c for c in ["ANNEE", "MOIS_NUM", "GRADE_PRINCIPAL"]
                     if c in base_agrege.columns]
    print(f"8b. Winsorisation REVENU_TOTAL "
          f"(P{int(SEUIL_WINS*100)}/P{int((1-SEUIL_WINS)*100)}) "
          f"— seuil groupe : {MIN_OBS_WINS} obs...")

    base_agrege = base_agrege.with_columns([
        pl.len().over(wins_full).alias("_n_citp"),
        pl.col("REVENU_TOTAL")
          .quantile(SEUIL_WINS, interpolation="nearest")
          .over(wins_full).alias("_pb_citp"),
        pl.col("REVENU_TOTAL")
          .quantile(1 - SEUIL_WINS, interpolation="nearest")
          .over(wins_full).alias("_ph_citp"),
        pl.col("REVENU_TOTAL")
          .quantile(SEUIL_WINS, interpolation="nearest")
          .over(wins_fallback).alias("_pb_grade"),
        pl.col("REVENU_TOTAL")
          .quantile(1 - SEUIL_WINS, interpolation="nearest")
          .over(wins_fallback).alias("_ph_grade"),
    ])

    n_obs_citp  = int((base_agrege["_n_citp"] >= MIN_OBS_WINS).sum())
    n_obs_grade = len(base_agrege) - n_obs_citp

    base_agrege = base_agrege.with_columns(
        pl.when(pl.col("_n_citp") >= MIN_OBS_WINS)
        .then(pl.col("REVENU_TOTAL").clip(pl.col("_pb_citp"), pl.col("_ph_citp")))
        .otherwise(pl.col("REVENU_TOTAL").clip(pl.col("_pb_grade"), pl.col("_ph_grade")))
        .alias("REVENU_TOTAL_W")
    ).drop(["_n_citp", "_pb_citp", "_ph_citp", "_pb_grade", "_ph_grade"])

    n_modif = int((base_agrege["REVENU_TOTAL_W"] != base_agrege["REVENU_TOTAL"]).sum())
    print(f"   Groupe mois × grade × CITP (≥{MIN_OBS_WINS} obs) : {n_obs_citp:,} obs")
    print(f"   Repli  mois × grade         (<{MIN_OBS_WINS} obs) : {n_obs_grade:,} obs")
    print(f"   Observations écrêtées : {n_modif:,}  ({100*n_modif/len(base_agrege):.2f}%)")
    print(f"   Gold inchangé — winsorisation en mémoire uniquement\n")
    value_col_indic = "REVENU_TOTAL_W"

# ============================================================
# 9. INDICATEURS SUR LE REVENU TOTAL
# ============================================================

print("9. Calcul indicateurs REVENU TOTAL...")

indic_grade = compute_indicators(
    base_agrege, ["ANNEE", "MOIS_NUM", "MOIS", "GRADE_PRINCIPAL"], value_col_indic
)
indic_grade_sexe = compute_indicators(
    base_agrege, ["ANNEE", "MOIS_NUM", "MOIS", "SEXE_STD", "GRADE_PRINCIPAL"], value_col_indic
)

if "Code_CITP" in base_agrege.columns:
    indic_citp = compute_indicators(
        base_agrege.filter(
            pl.col("Code_CITP").is_not_null() & (pl.col("Code_CITP") != "Non renseigné")
        ),
        ["ANNEE", "MOIS_NUM", "MOIS", "Code_CITP", "Metier_CITP"],
        value_col_indic,
    )
else:
    indic_citp = pl.DataFrame()

# Ventilations barème (si fuzzy join réussi en 7b)
indic_bareme       = pl.DataFrame()
indic_bareme_grade = pl.DataFrame()
indic_bareme_sexe  = pl.DataFrame()
global_brut        = None
stats_bareme_detail = pl.DataFrame()

if BAREME_DISPONIBLE:
    base_b = base_agrege.filter(pl.col("bareme").is_not_null())
    indic_bareme = compute_indicators(
        base_b, ["ANNEE", "MOIS_NUM", "MOIS", "bareme"], value_col_indic
    )
    indic_bareme_grade = compute_indicators(
        base_b, ["ANNEE", "MOIS_NUM", "MOIS", "bareme", "GRADE_PRINCIPAL"], value_col_indic
    )
    indic_bareme_sexe = compute_indicators(
        base_b, ["ANNEE", "MOIS_NUM", "MOIS", "bareme", "SEXE_STD"], value_col_indic
    )
    # Stats SALAIRE_BRUT ligne par ligne (pour MODELE_ANSTAT, calculé ici une seule fois)
    base_gv_b = base_gv.filter(
        pl.col("bareme_detail").is_not_null()
        & pl.col("SALAIRE_BRUT").is_not_null()
        & (pl.col("SALAIRE_BRUT") > 0)
    )
    if len(base_gv_b) > 0:
        global_brut = base_gv_b.select([
            pl.col("SALAIRE_BRUT").min().round(0).alias("salaire_brut_min"),
            pl.col("SALAIRE_BRUT").mean().round(0).alias("salaire_brut_moyen"),
            pl.col("SALAIRE_BRUT").max().round(0).alias("salaire_brut_max"),
            pl.col("SALAIRE_BRUT").median().round(0).alias("salaire_brut_mediane"),
        ]).row(0, named=True)
        stats_bareme_detail = (
            base_gv_b.group_by("bareme_detail")
            .agg([
                pl.col("SALAIRE_BRUT").min().round(0).alias("salaire_brut_min"),
                pl.col("SALAIRE_BRUT").mean().round(0).alias("salaire_brut_moyen"),
                pl.col("SALAIRE_BRUT").max().round(0).alias("salaire_brut_max"),
                pl.col("SALAIRE_BRUT").median().round(0).alias("salaire_brut_mediane"),
                pl.len().alias("nb_obs"),
            ])
            .sort("bareme_detail")
        )

total_indic = (len(indic_grade) + len(indic_grade_sexe) + len(indic_citp)
               + len(indic_bareme) + len(indic_bareme_grade) + len(indic_bareme_sexe))
print(f"   Grade            : {len(indic_grade):,} lignes")
print(f"   Grade × Sexe     : {len(indic_grade_sexe):,} lignes")
print(f"   CITP             : {len(indic_citp):,} lignes")
if BAREME_DISPONIBLE:
    print(f"   Barème           : {len(indic_bareme):,} lignes")
    print(f"   Barème × Grade   : {len(indic_bareme_grade):,} lignes")
    print(f"   Barème × Sexe    : {len(indic_bareme_sexe):,} lignes")
print(f"   Total            : {total_indic:,} lignes\n")

# ============================================================
# 10. INDICATEURS SALAIRE BRUT (par ligne, pivot grade)
# ============================================================

print("10. Calcul indicateurs SALAIRE BRUT (pivot grade)...")

base_gv_brut = base_gv
if SEUIL_WINS is not None:
    wins_full_brut     = [c for c in ["ANNEE", "MOIS_NUM", "GRADE", "Code_CITP"]
                          if c in base_gv.columns]
    wins_fallback_brut = [c for c in ["ANNEE", "MOIS_NUM", "GRADE"]
                          if c in base_gv.columns]

    base_gv_brut = base_gv.with_columns([
        pl.len().over(wins_full_brut).alias("_n_citp"),
        pl.col("SALAIRE_BRUT")
          .quantile(SEUIL_WINS, interpolation="nearest")
          .over(wins_full_brut).alias("_pb_citp"),
        pl.col("SALAIRE_BRUT")
          .quantile(1 - SEUIL_WINS, interpolation="nearest")
          .over(wins_full_brut).alias("_ph_citp"),
        pl.col("SALAIRE_BRUT")
          .quantile(SEUIL_WINS, interpolation="nearest")
          .over(wins_fallback_brut).alias("_pb_grade"),
        pl.col("SALAIRE_BRUT")
          .quantile(1 - SEUIL_WINS, interpolation="nearest")
          .over(wins_fallback_brut).alias("_ph_grade"),
    ])
    base_gv_brut = base_gv_brut.with_columns(
        pl.when(pl.col("_n_citp") >= MIN_OBS_WINS)
        .then(pl.col("SALAIRE_BRUT").clip(pl.col("_pb_citp"), pl.col("_ph_citp")))
        .otherwise(pl.col("SALAIRE_BRUT").clip(pl.col("_pb_grade"), pl.col("_ph_grade")))
        .alias("SALAIRE_BRUT")
    ).drop(["_n_citp", "_pb_citp", "_ph_citp", "_pb_grade", "_ph_grade"])

    n_modif_brut = int((base_gv_brut["SALAIRE_BRUT"] != base_gv["SALAIRE_BRUT"]).sum())
    print(f"   Winsorisation SALAIRE_BRUT (mois × grade × CITP / repli grade) : "
          f"{n_modif_brut:,} obs écrêtées")

brut_moyen = (
    base_gv_brut
    .group_by(["ANNEE", "MOIS_NUM", "MOIS", "GRADE"])
    .agg(pl.col("SALAIRE_BRUT").mean().round(0).alias("Salaire_brut_moyen"))
    .sort(["ANNEE", "MOIS_NUM"])
)
brut_median = (
    base_gv_brut
    .group_by(["ANNEE", "MOIS_NUM", "MOIS", "GRADE"])
    .agg(pl.col("SALAIRE_BRUT").median().round(0).alias("Salaire_brut_median"))
    .sort(["ANNEE", "MOIS_NUM"])
)

brut_moyen_wide = brut_moyen.pivot(
    values="Salaire_brut_moyen",
    index=["ANNEE", "MOIS_NUM", "MOIS"],
    on="GRADE"
).sort(["ANNEE", "MOIS_NUM"])

brut_median_wide = brut_median.pivot(
    values="Salaire_brut_median",
    index=["ANNEE", "MOIS_NUM", "MOIS"],
    on="GRADE"
).sort(["ANNEE", "MOIS_NUM"])

print(f"   Brut moyen  : {len(brut_moyen_wide)} lignes × {len(brut_moyen_wide.columns)} cols")
print(f"   Brut médian : {len(brut_median_wide)} lignes × {len(brut_median_wide.columns)} cols\n")

# ============================================================
# 11. MULTI-POSTES
# ============================================================

print("11. Création feuilles multi-postes...")

multi_detail = (
    base_agrege
    .group_by(["ANNEE", "MOIS_NUM", "MOIS", "GRADE_PRINCIPAL", "NB_POSTES"])
    .agg([
        pl.len().alias("Effectif"),
        pl.col("REVENU_TOTAL").mean().round(0).alias("Revenu_moyen"),
    ])
    .sort(["ANNEE", "MOIS_NUM", "GRADE_PRINCIPAL", "NB_POSTES"])
)

multi_detail = multi_detail.with_columns(
    pl.when(pl.col("NB_POSTES") >= 5)
    .then(pl.lit("5+"))
    .otherwise(pl.col("NB_POSTES").cast(pl.Utf8))
    .alias("NB_POSTES_LABEL")
)

multi_effectif = multi_detail.pivot(
    values="Effectif",
    index=["ANNEE", "MOIS_NUM", "MOIS", "GRADE_PRINCIPAL"],
    on="NB_POSTES_LABEL",
    aggregate_function="sum",
).sort(["ANNEE", "MOIS_NUM", "GRADE_PRINCIPAL"])

for c in ["1", "2", "3", "4", "5+"]:
    if c in multi_effectif.columns:
        multi_effectif = multi_effectif.with_columns(pl.col(c).fill_null(0))

distrib_postes = (
    base_agrege
    .group_by("NB_POSTES")
    .agg([
        pl.len().alias("Effectif"),
        (pl.len() * 100 / len(base_agrege)).round(2).alias("Pct"),
        pl.col("REVENU_TOTAL").mean().round(0).alias("Revenu_moyen"),
    ])
    .sort("NB_POSTES")
)

print(f"   Multi-postes détail  : {len(multi_effectif)} lignes")
print(f"   Distribution globale : {len(distrib_postes)} lignes\n")

# ============================================================
# 12. EXPORT EXCEL
# ============================================================

print("12. Export Excel...")

def appliquer_style_entete(ws, n_cols: int) -> None:
    fill  = PatternFill("solid", fgColor="4F81BD")
    font  = Font(color="FFFFFF", bold=True, size=11)
    align = Alignment(horizontal="center", vertical="center")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align


def ecrire_feuille(wb: Workbook, nom: str, df: pl.DataFrame) -> None:
    if len(df) == 0:
        return
    ws = wb.create_sheet(title=nom)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col)
    for i, row in enumerate(df.iter_rows(named=True), start=2):
        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=i, column=j, value=row[col])
    appliquer_style_entete(ws, len(df.columns))
    ws.freeze_panes = "A2"
    print(f"   ✓ {nom} ({len(df):,} lignes)")


wb = Workbook()
wb.remove(wb.active)

ecrire_feuille(wb, "REVENU_Grade_Detail",      indic_grade)
ecrire_feuille(wb, "REVENU_Grade_Sexe",         indic_grade_sexe)
if len(indic_citp) > 0:
    ecrire_feuille(wb, "REVENU_CITP_Detail",    indic_citp)
ecrire_feuille(wb, "SALAIRE_BRUT_Moyen",        brut_moyen_wide)
ecrire_feuille(wb, "SALAIRE_BRUT_Median",       brut_median_wide)
ecrire_feuille(wb, "MULTI_Par_Grade_NbPostes",  multi_effectif)
ecrire_feuille(wb, "MULTI_Distribution",        distrib_postes)

# --- Feuilles barème (si disponibles) ---
if BAREME_DISPONIBLE:
    ecrire_feuille(wb, "REVENU_Bareme_Detail",  indic_bareme)
    ecrire_feuille(wb, "REVENU_Bareme_Grade",   indic_bareme_grade)
    ecrire_feuille(wb, "REVENU_Bareme_Sexe",    indic_bareme_sexe)

    # --- Feuille MODELE_ANSTAT ---
    # Format identique à l'EXEMPLE_STATISTIQUE du fichier ANSTAT :
    # lignes = secteurs / sous-groupes, colonnes = min / moyen / max salaire brut
    if global_brut is not None and len(stats_bareme_detail) > 0:
        ws_anstat = wb.create_sheet("MODELE_ANSTAT")
        ws_anstat.sheet_view.showGridLines = False

        fill_titre   = PatternFill("solid", fgColor="1F4E79")
        fill_entete  = PatternFill("solid", fgColor="2E75B6")
        fill_global  = PatternFill("solid", fgColor="FFF2CC")
        fill_secteur = PatternFill("solid", fgColor="D6E4F0")
        fill_alt     = [PatternFill("solid", fgColor="EBF3FB"),
                        PatternFill("solid", fgColor="FFFFFF")]

        font_titre   = Font(bold=True, color="FFFFFF", size=12)
        font_entete  = Font(bold=True, color="FFFFFF", size=10)
        font_global  = Font(bold=True, color="7F6000", size=11)
        font_secteur = Font(bold=True, color="1F4E79", size=10)
        font_sous    = Font(size=10, italic=True)
        font_montant = Font(size=10)

        al_centre  = Alignment(horizontal="center", vertical="center")
        al_gauche  = Alignment(horizontal="left",   vertical="center")
        al_droite  = Alignment(horizontal="right",  vertical="center")
        al_indent  = Alignment(horizontal="left",   vertical="center", indent=2)

        annee_max_label = ""
        if "ANNEE" in base.columns:
            annees = base["ANNEE"].drop_nulls().unique().to_list()
            if annees:
                annee_max_label = str(sorted(annees)[-1])

        # Titre
        ws_anstat.merge_cells("A1:D1")
        titre_cell = ws_anstat.cell(
            row=1, column=1,
            value=f"MODÈLE DE DONNÉES DE SALAIRE DES FONCTIONNAIRES"
                  f"{' POUR ' + annee_max_label if annee_max_label else ''}"
        )
        titre_cell.font      = font_titre
        titre_cell.fill      = fill_titre
        titre_cell.alignment = al_centre
        ws_anstat.row_dimensions[1].height = 28

        # En-têtes colonnes
        ws_anstat.cell(row=2, column=1, value="salaires en FCFA").font = Font(size=10)
        for col_i, label in enumerate(
            ["salaire brut minimum", "salaire brut moyen", "salaire brut maximum"],
            start=2
        ):
            c = ws_anstat.cell(row=2, column=col_i, value=label)
            c.font      = font_entete
            c.fill      = fill_entete
            c.alignment = al_centre
        ws_anstat.row_dimensions[2].height = 22

        # Lookup dans stats_bareme_detail
        def get_val_anstat(key: str, col: str):
            if key == "_GLOBAL_":
                return global_brut.get(col)
            rows = stats_bareme_detail.filter(pl.col("bareme_detail") == key)
            if len(rows) == 0:
                return None
            return rows.row(0, named=True).get(col)

        row_anstat = 3
        sous_idx   = 0
        for libelle, bareme_key in ORDRE_MODELE_ANSTAT:
            est_global   = bareme_key == "_GLOBAL_"
            est_titre    = bareme_key is None
            est_sous_gpe = not est_global and not est_titre

            if est_global:
                for col_i, val in enumerate([
                    "TOUS SECTEURS CONFONDUS",
                    get_val_anstat("_GLOBAL_", "salaire_brut_min"),
                    get_val_anstat("_GLOBAL_", "salaire_brut_moyen"),
                    get_val_anstat("_GLOBAL_", "salaire_brut_max"),
                ], start=1):
                    c = ws_anstat.cell(row=row_anstat, column=col_i, value=val)
                    c.font      = font_global
                    c.fill      = fill_global
                    c.alignment = al_gauche if col_i == 1 else al_centre
                    if col_i > 1 and isinstance(val, (int, float)):
                        c.number_format = "#,##0"
                ws_anstat.row_dimensions[row_anstat].height = 22

            elif est_titre:
                ws_anstat.merge_cells(
                    start_row=row_anstat, start_column=1,
                    end_row=row_anstat,   end_column=4
                )
                c = ws_anstat.cell(row=row_anstat, column=1, value=libelle)
                c.font      = font_secteur
                c.fill      = fill_secteur
                c.alignment = al_gauche
                ws_anstat.row_dimensions[row_anstat].height = 20

            else:
                bg = fill_alt[sous_idx % 2]
                sous_idx += 1
                c1 = ws_anstat.cell(row=row_anstat, column=1, value=libelle)
                c1.font      = font_sous
                c1.fill      = bg
                c1.alignment = al_indent
                for col_i, stat_key in enumerate(
                    ["salaire_brut_min", "salaire_brut_moyen", "salaire_brut_max"],
                    start=2
                ):
                    val = get_val_anstat(bareme_key, stat_key)
                    c = ws_anstat.cell(row=row_anstat, column=col_i, value=val)
                    c.font          = font_montant
                    c.fill          = bg
                    c.alignment     = al_droite
                    c.number_format = "#,##0"
                ws_anstat.row_dimensions[row_anstat].height = 18

            row_anstat += 1

        # Largeurs colonnes
        ws_anstat.column_dimensions["A"].width = 55
        for col_letter in ["B", "C", "D"]:
            ws_anstat.column_dimensions[col_letter].width = 22
        ws_anstat.freeze_panes = "A3"
        print(f"   ✓ MODELE_ANSTAT ({row_anstat - 3} lignes)")

buf_xlsx = io.BytesIO()
wb.save(buf_xlsx)
buf_xlsx.seek(0)
xlsx_bytes = buf_xlsx.getvalue()

s3.put_object(
    Bucket=BUCKET_STAGING,
    Key=OUTPUT_KEY,
    Body=xlsx_bytes,
    ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

taille_kb = len(xlsx_bytes) / 1024
print(f"\n✅ Fichier créé : s3://{BUCKET_STAGING}/{OUTPUT_KEY}")
print(f"   Taille   : {taille_kb:.1f} KB")
print(f"   Feuilles : {len(wb.sheetnames)}\n")

print("=" * 70)
print("✓ CALCUL INDICATEURS TERMINÉ")
print("=" * 70)
print()
print("Prochaine étape : exécuter 08_creation_fichier_excel_avec_sommaire.py")